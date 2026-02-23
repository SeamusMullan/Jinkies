#!/usr/bin/env python3
"""Generate an Excel workbook with all Jinkies issues, counters, and a burndown chart."""

import json
import subprocess
from datetime import date

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, BarChart3D
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import DataBarRule, CellIsRule
from openpyxl.styles import (
    Alignment, Border, Font, NamedStyle, PatternFill, Side, numbers,
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Fetch issues
# ---------------------------------------------------------------------------
result = subprocess.run(
    ["gh", "issue", "list", "--repo", "SeamusMullan/Jinkies", "--state", "all",
     "--limit", "200", "--json", "number,title,labels,state,createdAt,closedAt"],
    capture_output=True, text=True, check=True,
)
issues = json.loads(result.stdout)

priority_order = {"P0": 0, "P1": 1, "P2": 2, "P3": 3}

rows = []
for issue in issues:
    labels = [l["name"] for l in issue.get("labels", [])]
    priority = next((l for l in labels if l.startswith("P")), "none")
    other_labels = ", ".join(l for l in labels if not l.startswith("P"))
    rows.append({
        "number": issue["number"],
        "title": issue["title"],
        "priority": priority,
        "labels": other_labels,
        "state": issue["state"],
        "created": issue["createdAt"][:10],
        "closed": issue["closedAt"][:10] if issue.get("closedAt") else "",
    })

rows.sort(key=lambda r: (priority_order.get(r["priority"], 99), r["number"]))

# ---------------------------------------------------------------------------
# Workbook setup
# ---------------------------------------------------------------------------
wb = Workbook()

# Colors
DARK_BG = PatternFill("solid", fgColor="1E1E2E")
HEADER_BG = PatternFill("solid", fgColor="2D2D44")
P1_FILL = PatternFill("solid", fgColor="FF4C4C")
P2_FILL = PatternFill("solid", fgColor="FFA500")
P3_FILL = PatternFill("solid", fgColor="6C757D")
OPEN_FILL = PatternFill("solid", fgColor="2D4A2D")
CLOSED_FILL = PatternFill("solid", fgColor="4A2D2D")
WHITE_FONT = Font(color="FFFFFF", size=10)
BOLD_WHITE = Font(color="FFFFFF", size=10, bold=True)
TITLE_FONT = Font(color="FFFFFF", size=14, bold=True)
SUBTITLE_FONT = Font(color="AAAAAA", size=10)
COUNTER_FONT = Font(color="FFFFFF", size=22, bold=True)
COUNTER_LABEL = Font(color="AAAAAA", size=9)
THIN_BORDER = Border(
    left=Side(style="thin", color="444466"),
    right=Side(style="thin", color="444466"),
    top=Side(style="thin", color="444466"),
    bottom=Side(style="thin", color="444466"),
)

priority_fills = {"P0": PatternFill("solid", fgColor="DC3545"), "P1": P1_FILL, "P2": P2_FILL, "P3": P3_FILL}


def style_range(ws, row, col_start, col_end, fill=None, font=None, border=None, alignment=None):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        if fill:
            cell.fill = fill
        if font:
            cell.font = font
        if border:
            cell.border = border
        if alignment:
            cell.alignment = alignment


# =========================================================================
# SHEET 1: Issues
# =========================================================================
ws = wb.active
ws.title = "Issues"
ws.sheet_properties.tabColor = "4C9AFF"

# Dark background for all visible cells
for r in range(1, len(rows) + 20):
    for c in range(1, 10):
        ws.cell(row=r, column=c).fill = DARK_BG

# Title
ws.merge_cells("B2:H2")
ws.cell(row=2, column=2, value="Jinkies - Issue Tracker").font = TITLE_FONT
ws.cell(row=2, column=2).fill = DARK_BG
ws.merge_cells("B3:H3")
ws.cell(row=3, column=2, value=f"Generated {date.today().isoformat()}  |  {len(rows)} total issues").font = SUBTITLE_FONT
ws.cell(row=3, column=2).fill = DARK_BG

# Counters row
open_count = sum(1 for r in rows if r["state"] == "OPEN")
closed_count = sum(1 for r in rows if r["state"] == "CLOSED")
p1_open = sum(1 for r in rows if r["priority"] == "P1" and r["state"] == "OPEN")
p2_open = sum(1 for r in rows if r["priority"] == "P2" and r["state"] == "OPEN")
p3_open = sum(1 for r in rows if r["priority"] == "P3" and r["state"] == "OPEN")

counter_defs = [
    ("B", "Open", open_count, "2D4A2D", "66BB6A"),
    ("C", "Closed", closed_count, "4A2D2D", "EF5350"),
    ("D", "P1 Open", p1_open, "4A1A1A", "FF4C4C"),
    ("E", "P2 Open", p2_open, "4A3A1A", "FFA500"),
    ("F", "P3 Open", p3_open, "2D2D3A", "6C757D"),
]

for i, (col_letter, label, value, bg_color, font_color) in enumerate(counter_defs):
    col = 2 + i
    cell_val = ws.cell(row=5, column=col, value=value)
    cell_val.font = Font(color=font_color, size=22, bold=True)
    cell_val.fill = PatternFill("solid", fgColor=bg_color)
    cell_val.alignment = Alignment(horizontal="center")
    cell_val.border = THIN_BORDER

    cell_lbl = ws.cell(row=6, column=col, value=label)
    cell_lbl.font = Font(color="AAAAAA", size=9)
    cell_lbl.fill = PatternFill("solid", fgColor=bg_color)
    cell_lbl.alignment = Alignment(horizontal="center")
    cell_lbl.border = THIN_BORDER

# Progress bar (percentage complete)
ws.merge_cells("G5:H5")
pct = closed_count / len(rows) * 100 if rows else 0
ws.cell(row=5, column=7, value=pct / 100).font = Font(color="4C9AFF", size=22, bold=True)
ws.cell(row=5, column=7).number_format = '0.0%'
ws.cell(row=5, column=7).fill = PatternFill("solid", fgColor="1A2A4A")
ws.cell(row=5, column=7).alignment = Alignment(horizontal="center")
ws.cell(row=5, column=7).border = THIN_BORDER
ws.merge_cells("G6:H6")
ws.cell(row=6, column=7, value="Completion").font = Font(color="AAAAAA", size=9)
ws.cell(row=6, column=7).fill = PatternFill("solid", fgColor="1A2A4A")
ws.cell(row=6, column=7).alignment = Alignment(horizontal="center")
ws.cell(row=6, column=7).border = THIN_BORDER

# Table headers
headers = ["#", "Priority", "Title", "Labels", "Status", "Created", "Closed"]
header_widths = [6, 10, 70, 30, 10, 12, 12]
header_row = 8
for i, (h, w) in enumerate(zip(headers, header_widths)):
    col = 2 + i
    cell = ws.cell(row=header_row, column=col, value=h)
    cell.font = BOLD_WHITE
    cell.fill = HEADER_BG
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="center" if i != 2 else "left")
    ws.column_dimensions[get_column_letter(col)].width = w

# Data rows
for idx, row in enumerate(rows):
    r = header_row + 1 + idx
    stripe = PatternFill("solid", fgColor="252540") if idx % 2 == 0 else DARK_BG

    # Number
    ws.cell(row=r, column=2, value=row["number"]).font = WHITE_FONT
    ws.cell(row=r, column=2).fill = stripe
    ws.cell(row=r, column=2).border = THIN_BORDER
    ws.cell(row=r, column=2).alignment = Alignment(horizontal="center")

    # Priority
    p = row["priority"]
    pfill = priority_fills.get(p, PatternFill("solid", fgColor="333355"))
    ws.cell(row=r, column=3, value=p).font = Font(color="FFFFFF", size=10, bold=True)
    ws.cell(row=r, column=3).fill = pfill
    ws.cell(row=r, column=3).border = THIN_BORDER
    ws.cell(row=r, column=3).alignment = Alignment(horizontal="center")

    # Title
    ws.cell(row=r, column=4, value=row["title"]).font = WHITE_FONT
    ws.cell(row=r, column=4).fill = stripe
    ws.cell(row=r, column=4).border = THIN_BORDER

    # Labels
    ws.cell(row=r, column=5, value=row["labels"]).font = Font(color="AAAAAA", size=9, italic=True)
    ws.cell(row=r, column=5).fill = stripe
    ws.cell(row=r, column=5).border = THIN_BORDER

    # Status
    is_open = row["state"] == "OPEN"
    status_fill = PatternFill("solid", fgColor="2D4A2D") if is_open else PatternFill("solid", fgColor="4A2D2D")
    status_font = Font(color="66BB6A" if is_open else "EF5350", size=10, bold=True)
    ws.cell(row=r, column=6, value=row["state"]).font = status_font
    ws.cell(row=r, column=6).fill = status_fill
    ws.cell(row=r, column=6).border = THIN_BORDER
    ws.cell(row=r, column=6).alignment = Alignment(horizontal="center")

    # Created
    ws.cell(row=r, column=7, value=row["created"]).font = Font(color="888888", size=9)
    ws.cell(row=r, column=7).fill = stripe
    ws.cell(row=r, column=7).border = THIN_BORDER
    ws.cell(row=r, column=7).alignment = Alignment(horizontal="center")

    # Closed
    ws.cell(row=r, column=8, value=row["closed"]).font = Font(color="888888", size=9)
    ws.cell(row=r, column=8).fill = stripe
    ws.cell(row=r, column=8).border = THIN_BORDER
    ws.cell(row=r, column=8).alignment = Alignment(horizontal="center")

# Freeze panes
ws.freeze_panes = f"B{header_row + 1}"

# =========================================================================
# SHEET 2: Burndown
# =========================================================================
ws2 = wb.create_sheet("Burndown")
ws2.sheet_properties.tabColor = "FF6B6B"

for r in range(1, 30):
    for c in range(1, 12):
        ws2.cell(row=r, column=c).fill = DARK_BG

ws2.merge_cells("B2:J2")
ws2.cell(row=2, column=2, value="Burndown Progress").font = TITLE_FONT
ws2.cell(row=2, column=2).fill = DARK_BG

# Build burndown data by priority
ws2.cell(row=4, column=2, value="Category").font = BOLD_WHITE
ws2.cell(row=4, column=2).fill = HEADER_BG
ws2.cell(row=4, column=2).border = THIN_BORDER
ws2.cell(row=4, column=3, value="Total").font = BOLD_WHITE
ws2.cell(row=4, column=3).fill = HEADER_BG
ws2.cell(row=4, column=3).border = THIN_BORDER
ws2.cell(row=4, column=4, value="Open").font = BOLD_WHITE
ws2.cell(row=4, column=4).fill = HEADER_BG
ws2.cell(row=4, column=4).border = THIN_BORDER
ws2.cell(row=4, column=5, value="Closed").font = BOLD_WHITE
ws2.cell(row=4, column=5).fill = HEADER_BG
ws2.cell(row=4, column=5).border = THIN_BORDER
ws2.cell(row=4, column=6, value="% Done").font = BOLD_WHITE
ws2.cell(row=4, column=6).fill = HEADER_BG
ws2.cell(row=4, column=6).border = THIN_BORDER

categories = [
    ("P1 - Critical", "P1"),
    ("P2 - High", "P2"),
    ("P3 - Low", "P3"),
    ("All Issues", None),
]

cat_fills = {
    "P1 - Critical": PatternFill("solid", fgColor="3A1A1A"),
    "P2 - High": PatternFill("solid", fgColor="3A2A1A"),
    "P3 - Low": PatternFill("solid", fgColor="2A2A3A"),
    "All Issues": PatternFill("solid", fgColor="1A2A3A"),
}

for i, (cat_name, prio) in enumerate(categories):
    r = 5 + i
    if prio:
        total = sum(1 for x in rows if x["priority"] == prio)
        op = sum(1 for x in rows if x["priority"] == prio and x["state"] == "OPEN")
        cl = sum(1 for x in rows if x["priority"] == prio and x["state"] == "CLOSED")
    else:
        total = len(rows)
        op = open_count
        cl = closed_count

    fill = cat_fills[cat_name]
    ws2.cell(row=r, column=2, value=cat_name).font = BOLD_WHITE
    ws2.cell(row=r, column=2).fill = fill
    ws2.cell(row=r, column=2).border = THIN_BORDER
    ws2.cell(row=r, column=3, value=total).font = WHITE_FONT
    ws2.cell(row=r, column=3).fill = fill
    ws2.cell(row=r, column=3).border = THIN_BORDER
    ws2.cell(row=r, column=3).alignment = Alignment(horizontal="center")
    ws2.cell(row=r, column=4, value=op).font = Font(color="66BB6A", size=10)
    ws2.cell(row=r, column=4).fill = fill
    ws2.cell(row=r, column=4).border = THIN_BORDER
    ws2.cell(row=r, column=4).alignment = Alignment(horizontal="center")
    ws2.cell(row=r, column=5, value=cl).font = Font(color="EF5350", size=10)
    ws2.cell(row=r, column=5).fill = fill
    ws2.cell(row=r, column=5).border = THIN_BORDER
    ws2.cell(row=r, column=5).alignment = Alignment(horizontal="center")
    ws2.cell(row=r, column=6, value=cl / total if total else 0).font = Font(color="4C9AFF", size=10, bold=True)
    ws2.cell(row=r, column=6).number_format = '0.0%'
    ws2.cell(row=r, column=6).fill = fill
    ws2.cell(row=r, column=6).border = THIN_BORDER
    ws2.cell(row=r, column=6).alignment = Alignment(horizontal="center")

ws2.column_dimensions["B"].width = 18
ws2.column_dimensions["C"].width = 10
ws2.column_dimensions["D"].width = 10
ws2.column_dimensions["E"].width = 10
ws2.column_dimensions["F"].width = 10

# Burndown bar chart - stacked open vs closed by priority
chart = BarChart()
chart.type = "col"
chart.grouping = "stacked"
chart.title = "Issue Burndown by Priority"
chart.y_axis.title = "Issues"
chart.x_axis.title = "Priority"
chart.style = 10
chart.width = 20
chart.height = 14

# Categories (P1, P2, P3, All)
cats = Reference(ws2, min_col=2, min_row=5, max_row=8)

# Open series
open_data = Reference(ws2, min_col=4, min_row=4, max_row=8)
chart.add_data(open_data, titles_from_data=True)
chart.series[0].graphicalProperties.solidFill = "66BB6A"

# Closed series
closed_data = Reference(ws2, min_col=5, min_row=4, max_row=8)
chart.add_data(closed_data, titles_from_data=True)
chart.series[1].graphicalProperties.solidFill = "EF5350"

chart.set_categories(cats)
chart.shape = 4
ws2.add_chart(chart, "B11")

# =========================================================================
# SHEET 3: By Label
# =========================================================================
ws3 = wb.create_sheet("By Label")
ws3.sheet_properties.tabColor = "66BB6A"

for r in range(1, 40):
    for c in range(1, 8):
        ws3.cell(row=r, column=c).fill = DARK_BG

ws3.merge_cells("B2:F2")
ws3.cell(row=2, column=2, value="Issues by Label").font = TITLE_FONT
ws3.cell(row=2, column=2).fill = DARK_BG

# Count issues per label
from collections import Counter
label_counter = Counter()
label_open = Counter()
for row in rows:
    for lbl in row["labels"].split(", "):
        lbl = lbl.strip()
        if lbl:
            label_counter[lbl] += 1
            if row["state"] == "OPEN":
                label_open[lbl] += 1

ws3.cell(row=4, column=2, value="Label").font = BOLD_WHITE
ws3.cell(row=4, column=2).fill = HEADER_BG
ws3.cell(row=4, column=2).border = THIN_BORDER
ws3.cell(row=4, column=3, value="Total").font = BOLD_WHITE
ws3.cell(row=4, column=3).fill = HEADER_BG
ws3.cell(row=4, column=3).border = THIN_BORDER
ws3.cell(row=4, column=4, value="Open").font = BOLD_WHITE
ws3.cell(row=4, column=4).fill = HEADER_BG
ws3.cell(row=4, column=4).border = THIN_BORDER
ws3.cell(row=4, column=5, value="Closed").font = BOLD_WHITE
ws3.cell(row=4, column=5).fill = HEADER_BG
ws3.cell(row=4, column=5).border = THIN_BORDER
ws3.cell(row=4, column=6, value="% Done").font = BOLD_WHITE
ws3.cell(row=4, column=6).fill = HEADER_BG
ws3.cell(row=4, column=6).border = THIN_BORDER

for i, (lbl, total) in enumerate(label_counter.most_common()):
    r = 5 + i
    op = label_open[lbl]
    cl = total - op
    stripe = PatternFill("solid", fgColor="252540") if i % 2 == 0 else DARK_BG

    ws3.cell(row=r, column=2, value=lbl).font = BOLD_WHITE
    ws3.cell(row=r, column=2).fill = stripe
    ws3.cell(row=r, column=2).border = THIN_BORDER
    ws3.cell(row=r, column=3, value=total).font = WHITE_FONT
    ws3.cell(row=r, column=3).fill = stripe
    ws3.cell(row=r, column=3).border = THIN_BORDER
    ws3.cell(row=r, column=3).alignment = Alignment(horizontal="center")
    ws3.cell(row=r, column=4, value=op).font = Font(color="66BB6A", size=10)
    ws3.cell(row=r, column=4).fill = stripe
    ws3.cell(row=r, column=4).border = THIN_BORDER
    ws3.cell(row=r, column=4).alignment = Alignment(horizontal="center")
    ws3.cell(row=r, column=5, value=cl).font = Font(color="EF5350", size=10)
    ws3.cell(row=r, column=5).fill = stripe
    ws3.cell(row=r, column=5).border = THIN_BORDER
    ws3.cell(row=r, column=5).alignment = Alignment(horizontal="center")
    ws3.cell(row=r, column=6, value=cl / total if total else 0).font = Font(color="4C9AFF", size=10, bold=True)
    ws3.cell(row=r, column=6).number_format = '0.0%'
    ws3.cell(row=r, column=6).fill = stripe
    ws3.cell(row=r, column=6).border = THIN_BORDER
    ws3.cell(row=r, column=6).alignment = Alignment(horizontal="center")

ws3.column_dimensions["B"].width = 20
ws3.column_dimensions["C"].width = 10
ws3.column_dimensions["D"].width = 10
ws3.column_dimensions["E"].width = 10
ws3.column_dimensions["F"].width = 10

# Label bar chart
label_chart = BarChart()
label_chart.type = "bar"
label_chart.grouping = "stacked"
label_chart.title = "Issues by Label (Open vs Closed)"
label_chart.x_axis.title = "Count"
label_chart.style = 10
label_chart.width = 22
label_chart.height = 16

label_count = len(label_counter)
label_cats = Reference(ws3, min_col=2, min_row=5, max_row=4 + label_count)
label_open_ref = Reference(ws3, min_col=4, min_row=4, max_row=4 + label_count)
label_closed_ref = Reference(ws3, min_col=5, min_row=4, max_row=4 + label_count)

label_chart.add_data(label_open_ref, titles_from_data=True)
label_chart.series[0].graphicalProperties.solidFill = "66BB6A"
label_chart.add_data(label_closed_ref, titles_from_data=True)
label_chart.series[1].graphicalProperties.solidFill = "EF5350"
label_chart.set_categories(label_cats)

ws3.add_chart(label_chart, f"B{5 + label_count + 2}")

# =========================================================================
# Save
# =========================================================================
import os
outpath = os.path.join(os.path.dirname(os.path.abspath(__file__)), "jinkies-issues.xlsx")
wb.save(outpath)
print(f"Excel saved to {outpath}")
print(f"  - {len(rows)} issues ({open_count} open, {closed_count} closed)")
print(f"  - 3 sheets: Issues, Burndown, By Label")
